#!/usr/bin/env python3
"""Compile a TACATDP XLSForm workbook to ODK XForm XML.

The generated XML is a local artifact. Dataverse is updated only by the seed
script that reads this artifact and writes it to FormVersions.XFormXml.
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "docs/Revised_TACATDP impact evaluation_20260712.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "artifacts/xforms"
XML_NS = {
    "h": "http://www.w3.org/1999/xhtml",
    "xf": "http://www.w3.org/2002/xforms",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compile XLSForm to XForm XML with TACATDP metadata checks.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Path to the XLSForm .xlsx workbook.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Directory for generated XML and metadata.")
    parser.add_argument("--version", default=None, help="Override form version. Defaults to UTC YYYYMMDDHHMMSSmmm.")
    parser.add_argument("--skip-pyxform-validate", action="store_true", help="Pass --skip_validate to xls2xform.")
    return parser.parse_args()


def timestamp_version() -> str:
    now = datetime.now(timezone.utc)
    return now.strftime("%Y%m%d%H%M%S") + f"{now.microsecond // 1000:03d}"


def read_settings(workbook: Path) -> dict[str, str]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    if "settings" not in wb.sheetnames:
        raise SystemExit("XLSForm workbook is missing the settings sheet.")
    sheet = wb["settings"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("XLSForm settings sheet is empty.")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    values = rows[1] if len(rows) > 1 else ()
    settings: dict[str, str] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = values[index] if index < len(values) else None
        settings[header] = "" if value is None else str(value).strip()
    return settings


def count_sheet_rows(workbook: Path, sheet_name: str) -> int:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return 0
    sheet = wb[sheet_name]
    return max(sheet.max_row - 1, 0)


def prepare_workbook(source: Path, output_dir: Path, form_title: str, version: str) -> tuple[Path, list[str]]:
    wb = load_workbook(source)
    if "settings" not in wb.sheetnames or "survey" not in wb.sheetnames:
        raise SystemExit("XLSForm workbook must include settings and survey sheets.")

    settings = wb["settings"]
    settings_headers = {str(cell.value).strip(): index + 1 for index, cell in enumerate(settings[1]) if cell.value is not None}
    if "form_title" in settings_headers:
        settings.cell(row=2, column=settings_headers["form_title"]).value = form_title
    if "version" in settings_headers:
        settings.cell(row=2, column=settings_headers["version"]).value = version

    survey = wb["survey"]
    survey_headers = {str(cell.value).strip(): index + 1 for index, cell in enumerate(survey[1]) if cell.value is not None}
    name_col = survey_headers.get("name")
    if name_col is None:
        raise SystemExit("XLSForm survey sheet is missing the name column.")

    seen: dict[str, int] = defaultdict(int)
    changes: list[str] = []
    for row in range(2, survey.max_row + 1):
        cell = survey.cell(row=row, column=name_col)
        name = str(cell.value).strip() if cell.value is not None else ""
        if not name:
            continue
        seen[name] += 1
        if seen[name] == 1:
            continue
        new_name = f"{name}_{seen[name]}"
        cell.value = new_name
        changes.append(f"survey row {row}: renamed duplicate name {name} -> {new_name}")

    output = output_dir / f"{source.stem}.compiled-source-{version}.xlsx"
    wb.save(output)
    return output, changes


def compile_xform(workbook: Path, output_xml: Path, skip_validate: bool) -> None:
    command = [str(ROOT / ".venv/bin/xls2xform"), str(workbook), str(output_xml)]
    if skip_validate:
        command.insert(1, "--skip_validate")
    subprocess.run(command, cwd=ROOT, check=True)


def override_xform_metadata(xml_path: Path, form_title: str, version: str) -> None:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    title = root.find(".//h:title", XML_NS)
    if title is not None:
        title.text = form_title
    data = root.find(".//xf:instance/*", XML_NS)
    if data is None:
        raise SystemExit("Compiled XForm has no primary instance data node.")
    data.attrib["version"] = version
    tree.write(xml_path, encoding="utf-8", xml_declaration=True)


def validate_xml(xml_path: Path, expected_form_id: str, expected_version: str) -> dict[str, object]:
    text = xml_path.read_text(encoding="utf-8")
    root = ET.fromstring(text)
    data = root.find(".//xf:instance/*", XML_NS)
    if data is None:
        raise SystemExit("Compiled XForm has no primary instance data node.")
    form_id = data.attrib.get("id", "")
    version = data.attrib.get("version", "")
    if expected_form_id and form_id != expected_form_id:
        raise SystemExit(f"Compiled XForm form id mismatch: expected {expected_form_id}, got {form_id}")
    if expected_version and version != expected_version:
        raise SystemExit(f"Compiled XForm version mismatch: expected {expected_version}, got {version}")
    body = root.find(".//h:body", XML_NS)
    if body is None:
        raise SystemExit("Compiled XForm has no h:body.")
    refs: list[str] = []
    for element in body.iter():
        ref = element.attrib.get("ref")
        if ref and ref.startswith("/"):
            refs.append(ref)
    duplicates = sorted({ref for ref in refs if refs.count(ref) > 1})
    if duplicates:
        raise SystemExit(f"Compiled XForm has duplicate body refs rejected by ODK Web Forms: {', '.join(duplicates)}")
    return {
        "form_id": form_id,
        "version": version,
        "xml_bytes": len(text.encode("utf-8")),
        "body_ref_count": len(refs),
    }


def main() -> int:
    args = parse_args()
    workbook = Path(args.workbook).resolve()
    output_dir = Path(args.output_dir).resolve()
    if not workbook.exists():
        raise SystemExit(f"Workbook not found: {workbook}")
    settings = read_settings(workbook)
    version = args.version or timestamp_version()
    settings["version"] = version
    form_title = "TACATDP Impact Evaluation"
    form_id = settings.get("form_id") or "tacatdp_impact_evaluation"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_xml = output_dir / f"{form_id}-{version}.xml"
    metadata_json = output_dir / f"{form_id}-{version}.json"

    prepared_workbook, sanitation_changes = prepare_workbook(workbook, output_dir, form_title, version)
    compile_xform(prepared_workbook, output_xml, args.skip_pyxform_validate)
    override_xform_metadata(output_xml, form_title, version)
    validation = validate_xml(output_xml, form_id, version)
    metadata = {
        "workbook": str(workbook.relative_to(ROOT)),
        "compiled_source_workbook": str(prepared_workbook.relative_to(ROOT)),
        "xform_xml": str(output_xml.relative_to(ROOT)),
        "form_title": form_title,
        "form_id": form_id,
        "version": version,
        "instance_name": settings.get("instance_name", ""),
        "survey_rows": count_sheet_rows(workbook, "survey"),
        "choices_rows": count_sheet_rows(workbook, "choices"),
        "sanitation_changes": sanitation_changes,
        **validation,
    }
    metadata_json.write_text(json.dumps(metadata, indent=2) + "\n", encoding="utf-8")

    print("# XLSForm compile complete")
    print(f"Workbook: {metadata['workbook']}")
    print(f"Compiled source workbook: {metadata['compiled_source_workbook']}")
    print(f"XForm XML: {metadata['xform_xml']}")
    print(f"Metadata: {metadata_json.relative_to(ROOT)}")
    print(f"Form title: {metadata['form_title']}")
    print(f"Form id: {metadata['form_id']}")
    print(f"Version: {metadata['version']}")
    print(f"Instance name: {metadata['instance_name']}")
    print(f"Survey rows: {metadata['survey_rows']}")
    print(f"Choices rows: {metadata['choices_rows']}")
    print(f"XForm bytes: {metadata['xml_bytes']}")
    print(f"Body refs: {metadata['body_ref_count']}")
    if sanitation_changes:
        print("Sanitation changes:")
        for change in sanitation_changes:
            print(f"- {change}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

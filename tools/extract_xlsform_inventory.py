#!/usr/bin/env python3
"""Extract reviewable inventory files from the TACATDP XLSForm workbook."""

from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "docs" / "Revised_TACATDP impact evaluation_modified_groups.xlsx"
DOCS = ROOT / "docs"

FIELD_INVENTORY = DOCS / "xlsform-field-inventory.csv"
LOGIC_MAP = DOCS / "xlsform-logic-map.csv"
CHOICE_LISTS = DOCS / "xlsform-choice-lists.csv"
SUMMARY = DOCS / "xlsform-summary.md"

ODK_RUNTIME_TYPES = {
    "start",
    "end",
    "deviceid",
    "subscriberid",
    "simserial",
    "phonenumber",
    "username",
}

INPUT_TYPES = {
    "date",
    "decimal",
    "geopoint",
    "integer",
    "select_multiple",
    "select_one",
    "text",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def base_type(type_value: str) -> str:
    return clean(type_value).split()[0] if clean(type_value) else ""


def choice_list(type_value: str) -> str:
    parts = clean(type_value).split(maxsplit=1)
    if len(parts) == 2 and parts[0] in {"select_one", "select_multiple"}:
        return parts[1]
    return ""


def read_sheet(workbook: Path, sheet_name: str) -> list[dict[str, str]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, str]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {headers[index]: clean(value) for index, value in enumerate(row) if headers[index]}
        if any(item.values()):
            item["_row"] = str(row_number)
            rows.append(item)
    return rows


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    survey = read_sheet(WORKBOOK, "survey")
    choices = read_sheet(WORKBOOK, "choices")
    settings = read_sheet(WORKBOOK, "settings")

    field_rows: list[dict[str, str]] = []
    logic_rows: list[dict[str, str]] = []
    seen_input_names: set[str] = set()
    dropped_duplicate_names: list[str] = []

    for row in survey:
        type_value = row.get("type", "")
        kind = base_type(type_value)
        if kind not in INPUT_TYPES:
            continue

        name = row.get("name", "")
        if not name:
            continue
        if name in seen_input_names:
            dropped_duplicate_names.append(name)
            continue
        seen_input_names.add(name)

        field = {
            "survey_row": row["_row"],
            "type": type_value,
            "base_type": kind,
            "name": name,
            "label_en": row.get("label::English (en)", ""),
            "label_sw": row.get("label::Swahili (sw)", ""),
            "hint_en": row.get("hint::English (en)", ""),
            "required": row.get("required", ""),
            "required_message": row.get("required message", ""),
            "relevance": row.get("relevance", ""),
            "constraint": row.get("constraint", ""),
            "constraint_message": row.get("constraint message", ""),
            "default": row.get("default", ""),
            "appearance": row.get("appearance", ""),
            "choice_filter": row.get("choice_filter", ""),
            "choice_list": choice_list(type_value),
            "powerapps_control": "",
            "microsoft_lists_column": "",
            "microsoft_lists_type": "",
            "implementation_status": "input",
            "notes": "",
        }
        field_rows.append(field)

        if any(field[key] for key in ("required", "relevance", "constraint", "default", "choice_filter")):
            logic_rows.append(
                {
                    "survey_row": field["survey_row"],
                    "name": field["name"],
                    "type": field["type"],
                    "required": field["required"],
                    "required_message": field["required_message"],
                    "relevance": field["relevance"],
                    "constraint": field["constraint"],
                    "constraint_message": field["constraint_message"],
                    "default": field["default"],
                    "choice_filter": field["choice_filter"],
                }
            )

    by_list: dict[str, list[dict[str, str]]] = {}
    used_choice_lists = {row["choice_list"] for row in field_rows if row["choice_list"]}
    for row in choices:
        list_name = row.get("list_name", "")
        if list_name in used_choice_lists:
            by_list.setdefault(list_name, []).append(row)

    choice_rows = []
    for list_name, rows in sorted(by_list.items()):
        choice_rows.append(
            {
                "list_name": list_name,
                "choice_count": str(len(rows)),
                "sample_values": " | ".join(row.get("value", "") for row in rows[:10]),
                "sample_labels_en": " | ".join(row.get("label::English (en)", "") for row in rows[:10]),
                "filter_columns_present": ", ".join(
                    column
                    for column in ("zone", "region", "district", "ward", "crop_name", "technology")
                    if any(row.get(column, "") for row in rows)
                ),
            }
        )

    write_csv(
        FIELD_INVENTORY,
        [
            "survey_row",
            "type",
            "base_type",
            "name",
            "label_en",
            "label_sw",
            "hint_en",
            "required",
            "required_message",
            "relevance",
            "constraint",
            "constraint_message",
            "default",
            "appearance",
            "choice_filter",
            "choice_list",
            "powerapps_control",
            "microsoft_lists_column",
            "microsoft_lists_type",
            "implementation_status",
            "notes",
        ],
        field_rows,
    )
    write_csv(
        LOGIC_MAP,
        [
            "survey_row",
            "name",
            "type",
            "required",
            "required_message",
            "relevance",
            "constraint",
            "constraint_message",
            "default",
            "choice_filter",
        ],
        logic_rows,
    )
    write_csv(
        CHOICE_LISTS,
        ["list_name", "choice_count", "sample_values", "sample_labels_en", "filter_columns_present"],
        choice_rows,
    )

    type_counts = Counter(base_type(row.get("type", "")) for row in survey)
    input_type_counts = Counter(row["base_type"] for row in field_rows)
    settings_row = next((row for row in settings if row.get("form_title") or row.get("form_id")), {})
    SUMMARY.write_text(
        "\n".join(
            [
                "# XLSForm Summary",
                "",
                f"- Workbook: `{WORKBOOK.relative_to(ROOT)}`",
                f"- Form title: {settings_row.get('form_title', '')}",
                f"- Form ID: {settings_row.get('form_id', '')}",
                f"- Version: {settings_row.get('version', '')}",
                f"- Source survey rows: {len(survey)}",
                f"- Maintained input fields: {len(field_rows)}",
                f"- Excluded non-input rows: {len(survey) - len(field_rows)}",
                f"- Source choice rows: {len(choices)}",
                f"- Choice lists used by input fields: {len(by_list)}",
                f"- Input required rules: {sum(1 for row in field_rows if row.get('required'))}",
                f"- Input relevance rules: {sum(1 for row in field_rows if row.get('relevance'))}",
                f"- Input constraint rules: {sum(1 for row in field_rows if row.get('constraint'))}",
                f"- Source calculations excluded from input inventory: {sum(1 for row in survey if row.get('calculation'))}",
                f"- Input choice filters: {sum(1 for row in field_rows if row.get('choice_filter'))}",
                f"- Duplicate input field names dropped: {len(dropped_duplicate_names)}",
                "",
                "Notes, calculations, groups, and ODK runtime metadata are excluded from the maintained field inventory because they are not data-entry fields.",
                "",
                "## Survey type counts",
                "",
                *[f"- `{key}`: {count}" for key, count in sorted(type_counts.items())],
                "",
                "## Maintained input type counts",
                "",
                *[f"- `{key}`: {count}" for key, count in sorted(input_type_counts.items())],
                "",
                "## Generated inventory files",
                "",
                f"- `{FIELD_INVENTORY.relative_to(ROOT)}`",
                f"- `{LOGIC_MAP.relative_to(ROOT)}`",
                f"- `{CHOICE_LISTS.relative_to(ROOT)}`",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"Wrote {FIELD_INVENTORY.relative_to(ROOT)} ({len(field_rows)} rows)")
    print(f"Wrote {LOGIC_MAP.relative_to(ROOT)} ({len(logic_rows)} rows)")
    print(f"Wrote {CHOICE_LISTS.relative_to(ROOT)} ({len(choice_rows)} rows)")
    print(f"Wrote {SUMMARY.relative_to(ROOT)}")


if __name__ == "__main__":
    main()

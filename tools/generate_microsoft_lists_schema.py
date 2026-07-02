#!/usr/bin/env python3
"""Generate Microsoft Lists schema artifacts from the TACATDP XLSForm inventory."""

from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
SCHEMAS = ROOT / "schemas"
TEMPLATES = SCHEMAS / "import-templates"
REFERENCE_DATA = SCHEMAS / "reference-data"
SCRIPTS = ROOT / "scripts"

WORKBOOK = DOCS / "Revised_TACATDP impact evaluation_modified_groups.xlsx"
FIELD_INVENTORY = DOCS / "xlsform-field-inventory.csv"

SCHEMA_JSON = SCHEMAS / "sharepoint-lists-schema.json"
FIELDS_CSV = SCHEMAS / "sharepoint-fields.csv"
MAPPING_CSV = SCHEMAS / "xlsform-to-list-mapping.csv"
SUMMARY_MD = SCHEMAS / "schema-summary.md"
CREATE_SCRIPT = SCRIPTS / "create-microsoft-lists.ps1"
IMPORT_REFERENCE_SCRIPT = SCRIPTS / "import-reference-data.ps1"
CREATE_CMD = SCRIPTS / "create-microsoft-lists.cmd"
IMPORT_REFERENCE_CMD = SCRIPTS / "import-reference-data.cmd"

LOCATION_LISTS = {"region", "district", "ward", "village"}
REFERENCE_SELECT_LISTS = {"branch", "crop_name", "cost_item", "unit_measure", *LOCATION_LISTS}

SECTION_BY_ROW = [
    (0, 28, "TACATDP_Profile"),
    (29, 39, "TACATDP_Agriculture"),
    (40, 55, "TACATDP_ResourceEfficiency"),
    (56, 87, "TACATDP_SocialInclusion"),
    (88, 333, "TACATDP_Beneficiaries"),
    (334, 351, "TACATDP_SafeguardsClimate"),
    (352, 360, "TACATDP_InsuranceGuarantee"),
    (361, 417, "TACATDP_GHGWaterYield"),
    (418, 9999, "TACATDP_ProductionIncome"),
]

SECTION_LISTS = [
    "TACATDP_Profile",
    "TACATDP_Agriculture",
    "TACATDP_ResourceEfficiency",
    "TACATDP_SocialInclusion",
    "TACATDP_Beneficiaries",
    "TACATDP_SafeguardsClimate",
    "TACATDP_InsuranceGuarantee",
    "TACATDP_GHGWaterYield",
    "TACATDP_ProductionIncome",
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def pascal_case(value: str) -> str:
    parts = re.split(r"[^A-Za-z0-9]+", value)
    return "".join(part[:1].upper() + part[1:] for part in parts if part)


def section_for_row(row_number: int) -> str:
    for start, end, section in SECTION_BY_ROW:
        if start <= row_number <= end:
            return section
    raise ValueError(f"No section mapping for row {row_number}")


def sharepoint_type(field: dict[str, str]) -> str:
    base_type = field["base_type"]
    name = field["name"].lower()
    label = field["label_en"].lower()
    if base_type == "integer":
        return "Number"
    if base_type == "decimal":
        if "tzs" in name or "cost" in name or "price" in name or "revenue" in name or "income" in name:
            return "Currency"
        if "tzs" in label or "cost" in label or "price" in label or "revenue" in label or "income" in label:
            return "Currency"
        return "Number"
    if base_type == "date":
        return "DateTime"
    return "Text"


def number_decimals(field: dict[str, str]) -> int | None:
    if field["base_type"] == "integer":
        return 0
    if field["base_type"] == "decimal":
        return 2
    return None


def field_def(
    internal_name: str,
    display_name: str | None = None,
    field_type: str = "Text",
    *,
    required: bool = False,
    indexed: bool = False,
    unique: bool = False,
    decimals: int | None = None,
    source: str = "",
    description: str = "",
) -> dict[str, Any]:
    item: dict[str, Any] = {
        "internalName": internal_name,
        "displayName": display_name or internal_name,
        "type": field_type,
        "required": required,
        "indexed": indexed,
        "unique": unique,
        "source": source,
        "description": description,
    }
    if decimals is not None:
        item["decimals"] = decimals
    return item


def system_section_fields() -> list[dict[str, Any]]:
    return [
        field_def("SubmissionKey", "Submission Key", required=True, indexed=True),
        field_def("FormVersion", "Form Version", indexed=True),
        field_def("SectionStatus", "Section Status", indexed=True),
        field_def("LastSavedAt", "Last Saved At", "DateTime"),
    ]


def make_list(title: str, description: str, fields: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "title": title,
        "template": "GenericList",
        "description": description,
        "fields": fields,
    }


def load_choices() -> list[dict[str, str]]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["choices"]
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, str]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {headers[index]: clean(value) for index, value in enumerate(row) if headers[index]}
        if any(item.values()):
            item["_row"] = str(row_number)
            rows.append(item)
    return rows


def choice_counts(choices: list[dict[str, str]]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for row in choices:
        counts[row["list_name"]] += 1
    return dict(counts)


def is_cost_line_field(name: str) -> bool:
    return bool(re.fullmatch(r"vc\d+_(cost_item|unit|quantity|unit_cost)", name))


def add_scalar_field(
    fields: list[dict[str, Any]],
    mapping: list[dict[str, Any]],
    field: dict[str, str],
    target_list: str,
    counts: dict[str, int],
) -> None:
    name = field["name"]
    base_type = field["base_type"]
    required_in_app = field["required"] == "yes"
    source = f"xlsform:{name}"

    if base_type == "geopoint":
        geo_fields = [
            field_def(f"{name}Raw", f"{field['label_en']} Raw", "Text", source=source),
            field_def(f"{name}Latitude", f"{field['label_en']} Latitude", "Number", decimals=8, source=source),
            field_def(f"{name}Longitude", f"{field['label_en']} Longitude", "Number", decimals=8, source=source),
            field_def(f"{name}Accuracy", f"{field['label_en']} Accuracy", "Number", decimals=2, source=source),
        ]
        fields.extend(geo_fields)
        target_columns = ";".join(item["internalName"] for item in geo_fields)
        mapping.append(mapping_row(field, target_list, target_columns, "Geopoint split into raw/latitude/longitude/accuracy"))
        return

    if base_type == "select_one":
        value_col = f"{name}Value"
        label_col = f"{name}Label"
        fields.append(field_def(value_col, f"{field['label_en']} Value", "Text", indexed=True, source=source))
        fields.append(field_def(label_col, f"{field['label_en']} Label", "Text", source=source))
        mapping.append(
            mapping_row(
                field,
                target_list,
                f"{value_col};{label_col}",
                "select_one stored as text value/label from reference data for delegation and export fidelity",
                required_in_app,
            )
        )
        return

    fields.append(
        field_def(
            name,
            field["label_en"] or pascal_case(name),
            sharepoint_type(field),
            decimals=number_decimals(field),
            source=source,
        )
    )
    mapping.append(mapping_row(field, target_list, name, "scalar input field", required_in_app))


def mapping_row(
    field: dict[str, str],
    target_list: str,
    target_columns: str,
    strategy: str,
    required_in_app: bool | None = None,
) -> dict[str, Any]:
    if required_in_app is None:
        required_in_app = field["required"] == "yes"
    return {
        "survey_row": field["survey_row"],
        "xlsform_name": field["name"],
        "xlsform_type": field["type"],
        "base_type": field["base_type"],
        "label_en": field["label_en"],
        "choice_list": field["choice_list"],
        "target_list": target_list,
        "target_columns": target_columns,
        "required_in_app": "yes" if required_in_app else "no",
        "required_in_sharepoint": "no",
        "relevance": field["relevance"],
        "constraint": field["constraint"],
        "choice_filter": field["choice_filter"],
        "storage_strategy": strategy,
    }


def generate_reference_data(choices: list[dict[str, str]]) -> None:
    REFERENCE_DATA.mkdir(parents=True, exist_ok=True)

    def rows_for(list_name: str) -> list[dict[str, str]]:
        return [row for row in choices if row["list_name"] == list_name]

    reference_specs = {
        "TACATDP_RefRegions.csv": (
            ["ChoiceValue", "LabelEn", "LabelSw", "SortOrder"],
            [
                {
                    "ChoiceValue": row["value"],
                    "LabelEn": row["label::English (en)"],
                    "LabelSw": row["label::Swahili (sw)"],
                    "SortOrder": index,
                }
                for index, row in enumerate(rows_for("region"), start=1)
            ],
        ),
        "TACATDP_RefDistricts.csv": (
            ["ChoiceValue", "LabelEn", "LabelSw", "RegionValue", "SortOrder"],
            [
                {
                    "ChoiceValue": row["value"],
                    "LabelEn": row["label::English (en)"],
                    "LabelSw": row["label::Swahili (sw)"],
                    "RegionValue": row["region"],
                    "SortOrder": index,
                }
                for index, row in enumerate(rows_for("district"), start=1)
            ],
        ),
        "TACATDP_RefWards.csv": (
            ["ChoiceValue", "LabelEn", "LabelSw", "RegionValue", "DistrictValue", "SortOrder"],
            [
                {
                    "ChoiceValue": row["value"],
                    "LabelEn": row["label::English (en)"],
                    "LabelSw": row["label::Swahili (sw)"],
                    "RegionValue": row["region"],
                    "DistrictValue": row["district"],
                    "SortOrder": index,
                }
                for index, row in enumerate(rows_for("ward"), start=1)
            ],
        ),
        "TACATDP_RefVillages.csv": (
            ["ChoiceValue", "LabelEn", "LabelSw", "RegionValue", "DistrictValue", "WardValue", "SortOrder"],
            [
                {
                    "ChoiceValue": row["value"],
                    "LabelEn": row["label::English (en)"],
                    "LabelSw": row["label::Swahili (sw)"],
                    "RegionValue": row["region"],
                    "DistrictValue": row["district"],
                    "WardValue": row["ward"],
                    "SortOrder": index,
                }
                for index, row in enumerate(rows_for("village"), start=1)
            ],
        ),
        "TACATDP_RefBranches.csv": (
            ["ChoiceValue", "LabelEn", "LabelSw", "ZoneValue", "SortOrder"],
            [
                {
                    "ChoiceValue": row["value"],
                    "LabelEn": row["label::English (en)"],
                    "LabelSw": row["label::Swahili (sw)"],
                    "ZoneValue": row["zone"],
                    "SortOrder": index,
                }
                for index, row in enumerate(rows_for("branch"), start=1)
            ],
        ),
    }

    used_ref_lists = set(REFERENCE_SELECT_LISTS) | {
        "crop_adaptation",
        "climate_risk",
        "adaptation",
        "additional",
        "SDGs",
    }
    generic_rows = []
    for row in choices:
        list_name = row["list_name"]
        if list_name in LOCATION_LISTS or list_name == "branch":
            continue
        if list_name not in used_ref_lists and int(row.get("_row", "0")):
            pass
        generic_rows.append(
            {
                "ChoiceListName": list_name,
                "ChoiceValue": row["value"],
                "LabelEn": row["label::English (en)"],
                "LabelSw": row["label::Swahili (sw)"],
                "ZoneValue": row.get("zone", ""),
                "RegionValue": row.get("region", ""),
                "DistrictValue": row.get("district", ""),
                "WardValue": row.get("ward", ""),
                "CropNameValue": row.get("crop_name", ""),
                "TechnologyValue": row.get("technology", ""),
            }
        )
    reference_specs["TACATDP_RefChoices.csv"] = (
        [
            "ChoiceListName",
            "ChoiceValue",
            "LabelEn",
            "LabelSw",
            "ZoneValue",
            "RegionValue",
            "DistrictValue",
            "WardValue",
            "CropNameValue",
            "TechnologyValue",
        ],
        generic_rows,
    )

    for filename, (headers, rows) in reference_specs.items():
        write_csv(REFERENCE_DATA / filename, headers, rows)


def generate_pnp_script() -> None:
    SCRIPTS.mkdir(parents=True, exist_ok=True)
    CREATE_SCRIPT.write_text(
        r'''param(
    [Parameter(Mandatory = $true)]
    [string]$SiteUrl,

    [string]$SchemaPath = "../schemas/sharepoint-lists-schema.json"
)

$ErrorActionPreference = "Stop"

function Ensure-PnPPowerShell {
    if ($PSVersionTable.PSEdition -eq "Desktop") {
        throw "PnP.PowerShell requires PowerShell 7+. Run this script with pwsh, or use create-microsoft-lists.cmd."
    }

    try {
        Import-Module PnP.PowerShell -ErrorAction Stop
        return
    } catch {
        Write-Host "PnP.PowerShell is not available in this PowerShell profile. Installing for CurrentUser..."
    }

    $gallery = Get-PSRepository -Name PSGallery -ErrorAction SilentlyContinue
    if ($gallery -and $gallery.InstallationPolicy -ne "Trusted") {
        Set-PSRepository -Name PSGallery -InstallationPolicy Trusted
    }

    Install-Module PnP.PowerShell -Scope CurrentUser -Force -AllowClobber
    Import-Module PnP.PowerShell -ErrorAction Stop
}

Ensure-PnPPowerShell

$resolvedSchemaPath = Resolve-Path $SchemaPath
$schema = Get-Content $resolvedSchemaPath -Raw | ConvertFrom-Json

Connect-PnPOnline -Url $SiteUrl -Interactive

foreach ($list in $schema.lists) {
    $existingList = Get-PnPList -Identity $list.title -ErrorAction SilentlyContinue
    if (-not $existingList) {
        Add-PnPList -Title $list.title -Template GenericList -OnQuickLaunch:$false | Out-Null
    }

    foreach ($field in $list.fields) {
        $existingField = Get-PnPField -List $list.title -Identity $field.internalName -ErrorAction SilentlyContinue
        if (-not $existingField) {
            $required = [bool]$field.required
            Add-PnPField `
                -List $list.title `
                -DisplayName $field.displayName `
                -InternalName $field.internalName `
                -Type $field.type `
                -Required:$required `
                -AddToDefaultView | Out-Null
        }

        $values = @{}
        if ($field.indexed) { $values["Indexed"] = $true }
        if ($field.unique) { $values["EnforceUniqueValues"] = $true }
        if ($null -ne $field.decimals) { $values["Decimals"] = [int]$field.decimals }
        if ($values.Count -gt 0) {
            Set-PnPField -List $list.title -Identity $field.internalName -Values $values | Out-Null
        }
    }
}
''',
        encoding="utf-8",
    )
    IMPORT_REFERENCE_SCRIPT.write_text(
        r'''param(
    [Parameter(Mandatory = $true)]
    [string]$SiteUrl,

    [string]$ReferenceDataPath = "../schemas/reference-data"
)

$ErrorActionPreference = "Stop"

function Ensure-PnPPowerShell {
    if ($PSVersionTable.PSEdition -eq "Desktop") {
        throw "PnP.PowerShell requires PowerShell 7+. Run this script with pwsh, or use import-reference-data.cmd."
    }

    try {
        Import-Module PnP.PowerShell -ErrorAction Stop
        return
    } catch {
        Write-Host "PnP.PowerShell is not available in this PowerShell profile. Installing for CurrentUser..."
    }

    $gallery = Get-PSRepository -Name PSGallery -ErrorAction SilentlyContinue
    if ($gallery -and $gallery.InstallationPolicy -ne "Trusted") {
        Set-PSRepository -Name PSGallery -InstallationPolicy Trusted
    }

    Install-Module PnP.PowerShell -Scope CurrentUser -Force -AllowClobber
    Import-Module PnP.PowerShell -ErrorAction Stop
}

Ensure-PnPPowerShell

$resolvedReferencePath = Resolve-Path $ReferenceDataPath
Connect-PnPOnline -Url $SiteUrl -Interactive

Get-ChildItem -Path $resolvedReferencePath -Filter "*.csv" | ForEach-Object {
    $listTitle = $_.BaseName
    $rows = Import-Csv $_.FullName
    Write-Host "Importing $($rows.Count) rows into $listTitle"

    foreach ($row in $rows) {
        $values = @{}
        foreach ($property in $row.PSObject.Properties) {
            if ($property.Value -ne $null -and $property.Value -ne "") {
                $values[$property.Name] = $property.Value
            }
        }
        if ($values.ContainsKey("ChoiceListName")) {
            $values["Title"] = "{0}-{1}" -f $values["ChoiceListName"], $values["ChoiceValue"]
        } elseif ($values.ContainsKey("ChoiceValue")) {
            $values["Title"] = $values["ChoiceValue"]
        }
        Add-PnPListItem -List $listTitle -Values $values | Out-Null
    }
}
''',
        encoding="utf-8",
    )
    CREATE_CMD.write_text(
        r'''@echo off
where pwsh >nul 2>nul
if errorlevel 1 (
  echo PowerShell 7+ ^(pwsh^) is required for PnP.PowerShell.
  echo Install it from https://aka.ms/powershell-release?tag=stable then rerun this command.
  exit /b 1
)
pwsh -NoProfile -ExecutionPolicy Bypass -File "%~dp0create-microsoft-lists.ps1" %*
''',
        encoding="utf-8",
    )
    IMPORT_REFERENCE_CMD.write_text(
        r'''@echo off
where pwsh >nul 2>nul
if errorlevel 1 (
  echo PowerShell 7+ ^(pwsh^) is required for PnP.PowerShell.
  echo Install it from https://aka.ms/powershell-release?tag=stable then rerun this command.
  exit /b 1
)
pwsh -NoProfile -ExecutionPolicy Bypass -File "%~dp0import-reference-data.ps1" %*
''',
        encoding="utf-8",
    )


def main() -> None:
    for path in (SCHEMAS, TEMPLATES, REFERENCE_DATA, SCRIPTS):
        path.mkdir(parents=True, exist_ok=True)

    fields = read_csv(FIELD_INVENTORY)
    choices = load_choices()
    counts = choice_counts(choices)

    list_fields: dict[str, list[dict[str, Any]]] = {
        "TACATDP_Submissions": [
            field_def("SubmissionKey", "Submission Key", required=True, indexed=True, unique=True),
            field_def("FormId", "Form ID", indexed=True),
            field_def("FormVersion", "Form Version", indexed=True),
            field_def("SubmissionStatus", "Submission Status", indexed=True),
            field_def("EnumeratorEmail", "Enumerator Email", indexed=True),
            field_def("EnumeratorName", "Enumerator Name"),
            field_def("StartedAt", "Started At", "DateTime"),
            field_def("SubmittedAt", "Submitted At", "DateTime"),
            field_def("LastSavedAt", "Last Saved At", "DateTime"),
            field_def("Customer_ID", "Customer ID", "Number", indexed=True, decimals=0),
            field_def("Customer_Name", "Customer Name"),
            field_def("RegionValue", "Region Value", indexed=True),
            field_def("DistrictValue", "District Value", indexed=True),
            field_def("WardValue", "Ward Value", indexed=True),
            field_def("VillageValue", "Village Value", indexed=True),
        ],
        "TACATDP_MultiSelectAnswers": [
            field_def("SubmissionKey", "Submission Key", required=True, indexed=True),
            field_def("SectionKey", "Section Key", indexed=True),
            field_def("QuestionName", "Question Name", required=True, indexed=True),
            field_def("ChoiceListName", "Choice List Name", required=True, indexed=True),
            field_def("ChoiceValue", "Choice Value", required=True, indexed=True),
            field_def("ChoiceLabelEn", "Choice Label English"),
            field_def("ChoiceLabelSw", "Choice Label Swahili"),
            field_def("SortOrder", "Sort Order", "Number", decimals=0),
        ],
        "TACATDP_ProductionCostLines": [
            field_def("SubmissionKey", "Submission Key", required=True, indexed=True),
            field_def("StageCode", "Stage Code", required=True, indexed=True),
            field_def("StageLabel", "Stage Label"),
            field_def("CostItemValue", "Cost Item Value", indexed=True),
            field_def("CostItemLabel", "Cost Item Label"),
            field_def("UnitValue", "Unit Value"),
            field_def("UnitLabel", "Unit Label"),
            field_def("Quantity", "Quantity", "Number", decimals=2),
            field_def("UnitCostTZS", "Unit Cost TZS", "Currency", decimals=2),
            field_def("LineTotalTZS", "Line Total TZS", "Currency", decimals=2),
        ],
        "TACATDP_RefRegions": [
            field_def("ChoiceValue", "Choice Value", required=True, indexed=True, unique=True),
            field_def("LabelEn", "Label English", required=True),
            field_def("LabelSw", "Label Swahili"),
            field_def("SortOrder", "Sort Order", "Number", decimals=0),
        ],
        "TACATDP_RefDistricts": [
            field_def("ChoiceValue", "Choice Value", required=True, indexed=True),
            field_def("LabelEn", "Label English", required=True),
            field_def("LabelSw", "Label Swahili"),
            field_def("RegionValue", "Region Value", required=True, indexed=True),
            field_def("SortOrder", "Sort Order", "Number", decimals=0),
        ],
        "TACATDP_RefWards": [
            field_def("ChoiceValue", "Choice Value", required=True, indexed=True),
            field_def("LabelEn", "Label English", required=True),
            field_def("LabelSw", "Label Swahili"),
            field_def("RegionValue", "Region Value", required=True, indexed=True),
            field_def("DistrictValue", "District Value", required=True, indexed=True),
            field_def("SortOrder", "Sort Order", "Number", decimals=0),
        ],
        "TACATDP_RefVillages": [
            field_def("ChoiceValue", "Choice Value", required=True, indexed=True),
            field_def("LabelEn", "Label English", required=True),
            field_def("LabelSw", "Label Swahili"),
            field_def("RegionValue", "Region Value", required=True, indexed=True),
            field_def("DistrictValue", "District Value", required=True, indexed=True),
            field_def("WardValue", "Ward Value", required=True, indexed=True),
            field_def("SortOrder", "Sort Order", "Number", decimals=0),
        ],
        "TACATDP_RefBranches": [
            field_def("ChoiceValue", "Choice Value", required=True, indexed=True),
            field_def("LabelEn", "Label English", required=True),
            field_def("LabelSw", "Label Swahili"),
            field_def("ZoneValue", "Zone Value", required=True, indexed=True),
            field_def("SortOrder", "Sort Order", "Number", decimals=0),
        ],
        "TACATDP_RefChoices": [
            field_def("ChoiceListName", "Choice List Name", required=True, indexed=True),
            field_def("ChoiceValue", "Choice Value", required=True, indexed=True),
            field_def("LabelEn", "Label English", required=True),
            field_def("LabelSw", "Label Swahili"),
            field_def("ZoneValue", "Zone Value", indexed=True),
            field_def("RegionValue", "Region Value", indexed=True),
            field_def("DistrictValue", "District Value", indexed=True),
            field_def("WardValue", "Ward Value", indexed=True),
            field_def("CropNameValue", "Crop Name Value", indexed=True),
            field_def("TechnologyValue", "Technology Value", indexed=True),
        ],
    }

    for section in SECTION_LISTS:
        list_fields[section] = system_section_fields()

    mapping: list[dict[str, Any]] = []
    for field in fields:
        name = field["name"]
        if field["base_type"] == "select_multiple":
            mapping.append(mapping_row(field, "TACATDP_MultiSelectAnswers", "QuestionName;ChoiceValue", "select_multiple child rows"))
            continue
        if is_cost_line_field(name):
            mapping.append(mapping_row(field, "TACATDP_ProductionCostLines", "StageCode;CostItemValue;UnitValue;Quantity;UnitCostTZS", "normalized production cost line"))
            continue
        target = section_for_row(int(field["survey_row"]))
        add_scalar_field(list_fields[target], mapping, field, target, counts)

    schema = {
        "version": "1.0",
        "sourceWorkbook": str(WORKBOOK.relative_to(ROOT)),
        "sourceInventory": str(FIELD_INVENTORY.relative_to(ROOT)),
        "lists": [
            make_list(title, f"TACATDP generated schema for {title}", fields)
            for title, fields in list_fields.items()
        ],
    }
    SCHEMA_JSON.write_text(json.dumps(schema, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    field_rows = []
    for list_def in schema["lists"]:
        for field in list_def["fields"]:
            field_rows.append(
                {
                    "list": list_def["title"],
                    "internalName": field["internalName"],
                    "displayName": field["displayName"],
                    "type": field["type"],
                    "required": field["required"],
                    "indexed": field["indexed"],
                    "unique": field["unique"],
                    "decimals": field.get("decimals", ""),
                    "source": field["source"],
                    "description": field["description"],
                }
            )
    write_csv(
        FIELDS_CSV,
        ["list", "internalName", "displayName", "type", "required", "indexed", "unique", "decimals", "source", "description"],
        field_rows,
    )
    write_csv(
        MAPPING_CSV,
        [
            "survey_row",
            "xlsform_name",
            "xlsform_type",
            "base_type",
            "label_en",
            "choice_list",
            "target_list",
            "target_columns",
            "required_in_app",
            "required_in_sharepoint",
            "relevance",
            "constraint",
            "choice_filter",
            "storage_strategy",
        ],
        mapping,
    )

    for list_def in schema["lists"]:
        headers = ["Title"] + [field["internalName"] for field in list_def["fields"]]
        write_csv(TEMPLATES / f"{list_def['title']}.csv", headers, [])

    generate_reference_data(choices)
    generate_pnp_script()

    SUMMARY_MD.write_text(
        "\n".join(
            [
                "# Microsoft Lists Import Schema Summary",
                "",
                f"- Lists defined: {len(schema['lists'])}",
                f"- SharePoint fields defined: {len(field_rows)}",
                f"- XLSForm input mappings: {len(mapping)}",
                "",
                "## Generated artifacts",
                "",
                f"- `{SCHEMA_JSON.relative_to(ROOT)}`",
                f"- `{FIELDS_CSV.relative_to(ROOT)}`",
                f"- `{MAPPING_CSV.relative_to(ROOT)}`",
                f"- `{TEMPLATES.relative_to(ROOT)}/`",
                f"- `{REFERENCE_DATA.relative_to(ROOT)}/`",
                f"- `{CREATE_SCRIPT.relative_to(ROOT)}`",
                f"- `{IMPORT_REFERENCE_SCRIPT.relative_to(ROOT)}`",
                f"- `{CREATE_CMD.relative_to(ROOT)}`",
                f"- `{IMPORT_REFERENCE_CMD.relative_to(ROOT)}`",
                "",
                "The PnP PowerShell scripts create lists with typed columns and import reference data. CSV templates are provided for submission-data import after list creation.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"Wrote {SCHEMA_JSON.relative_to(ROOT)}")
    print(f"Wrote {FIELDS_CSV.relative_to(ROOT)} ({len(field_rows)} fields)")
    print(f"Wrote {MAPPING_CSV.relative_to(ROOT)} ({len(mapping)} mappings)")
    print(f"Wrote {TEMPLATES.relative_to(ROOT)}/*.csv")
    print(f"Wrote {REFERENCE_DATA.relative_to(ROOT)}/*.csv")
    print(f"Wrote {CREATE_SCRIPT.relative_to(ROOT)}")
    print(f"Wrote {IMPORT_REFERENCE_SCRIPT.relative_to(ROOT)}")
    print(f"Wrote {CREATE_CMD.relative_to(ROOT)}")
    print(f"Wrote {IMPORT_REFERENCE_CMD.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
